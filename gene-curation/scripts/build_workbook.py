#!/usr/bin/env python3
"""Primary end-of-curation output: the GCI copy-paste workbook. Build it AFTER scoring/curation
and BEFORE the Step-9 cross-check (the cross-check then validates the rows in this file).

Every tab maps 1:1 to a GCI curation form so the curator can paste columns straight onto the
website. PASTE columns (navy header) map to GCI fields; REVIEW columns (grey header) are human-
readable extras for manual review (brief description, HPO labels, scores, rationale) — do NOT paste.

Fill CFG + the evidence lists (extend HPO_LABELS). Empty lists still emit a header-only tab.
Requires: pip install openpyxl      Usage: python3 build_workbook.py [out.xlsx]
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================== FILL THIS ==============================
CFG = dict(gene="GENE", disease="Disease name", mondo="MONDO:XXXXXXX",
           genetic_pts=12, experimental_pts=6, classification="Definitive")

HPO_LABELS = {"HP:0000253": "Progressive microcephaly", "HP:0025104": "Capillary malformation"}

INDIVIDUALS = [dict(
    label="Author Case1", brief="One-line plain-language clinical summary for review.",
    proband="Yes", hpo_ids="HP:0000253, HP:0025104", hpo_free="", not_hpo="", not_free="",
    sex="Male", country="", ethnicity="", race="", age_type="Report", age_val="", age_unit="",
    prev_testing="", genomewide="Yes", method1="Exome sequencing", method2="Sanger", geno_desc="",
    intrans="Yes", homozygous="No", var1="NM_XXXXXX.X:c.___ (p.___)", var1_clinvar="",
    var2="NM_XXXXXX.X:c.___ (p.___)", var2_clinvar="", addl="", pmid="XXXXXXXX",
    v1_score="0.1", v2_score="1.5", proband_score="1.5", rationale="missense+null=1.6 -> GCI 1.5")]

EXPERIMENTS = [dict(
    paper="Author 20XX", pmid="XXXXXXXX", etype="Model Systems", subtype="Non-human model organism",
    name="Gene-null mouse", go_id="", go_free="", func_evidence="", interact_gene="", interact_type="",
    detect_method="", implicated="", uberon_id="", uberon_free="", normally_expressed="",
    cell_type="", normal_func_go="", gene_alt="Constitutive knockout", altered_evidence="",
    model_hpo="HP:0000252", model_free="phenotype…", human_hpo="HP:0000253", human_free="phenotype…",
    rescue_method="", wt_rescues="", pt_var_rescues="", explanation="how model resembles disease",
    assoc_clinvar="", status="Score", points="2.0", category="Models & Rescue (max 4)")]

FAMILIES = []       # dict(label,mondo,hpo_ids,hpo_free,not_hpo,not_free,country,ethnicity,race,
                    #      prev_testing,method1,method2,geno_desc,n_affected,n_unaffected,n_seg,
                    #      inconsistent,consanguineous,pedigree,pub_lod,include_lod,proband,addl,pmid,elod,notes)
GROUPS = []         # dict(label,mondo,hpo_ids,hpo_free,not_hpo,not_free,n_male,n_female,country,
                    #      ethnicity,race,age_type,age_range,age_unit,total,n_fam,n_nofam,n_var,n_novar,
                    #      n_othergene,othergenes,prev_testing,genomewide,method1,method2,geno_desc,addl,pmid)
CASECONTROL = []    # see COLS_CC keys below
SEGREGATION = []    # dict(family,pub,pmid,affected,segregations,unaffected,elod,notes)
# =======================================================================

PASTE = PatternFill("solid", fgColor="1F4E78"); REVIEW = PatternFill("solid", fgColor="595959")
HF = Font(color="FFFFFF", bold=True, size=10); TOT = PatternFill("solid", fgColor="DDEBF7")
FLAG = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="BFBFBF"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical="top"); CTR = Alignment("center", "center", wrap_text=True)

def hpo_desc(ids): return "; ".join(HPO_LABELS.get(x.strip(), x.strip()) for x in (ids or "").split(",") if x.strip())
def width_for(h):
    k = h.lower()
    if any(w in k for w in ("description", "explanation", "rationale", "brief", "additional", "evidence for", "notes")): return 40
    if any(w in k for w in ("phenotype", "hpo", "descriptor", "free text", "variant", "label", "method", "gene alteration")): return 26
    return 15

def sheet(wb, name, columns, rows, note=None, freeze="B2"):
    """columns = list of (header, kind, key|callable). kind in {'paste','review'}."""
    ws = wb.create_sheet(name)
    ws.append([c[0] for c in columns])
    for i, (h, kind, _) in enumerate(columns, 1):
        c = ws.cell(1, i); c.fill = PASTE if kind == "paste" else REVIEW
        c.font = HF; c.alignment = CTR; c.border = B
    for r in rows:
        ws.append([(k(r) if callable(k) else r.get(k, "")) for (_, _, k) in columns])
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 1)):
        for c in row: c.border = B; c.alignment = WRAP
    for i, (h, _, _) in enumerate(columns, 1): ws.column_dimensions[get_column_letter(i)].width = width_for(h)
    ws.freeze_panes = freeze
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row): ws.row_dimensions[r[0].row].height = 60
    if note or not rows:
        n = ws.max_row + 2
        ws.cell(n, 1, value=(note or "(no rows — delete this tab if this evidence type is not used in the curation)")).font = Font(italic=True, color="7F7F7F")
    return ws

def build():
    wb = Workbook(); wb.remove(wb.active)

    # ---- Individual ----
    COLS_I = [
        ("Individual Label", "paste", "label"), ("Brief description (review)", "review", "brief"),
        ("Proband?", "paste", "proband"),
        ("Disease (MONDO — name)", "paste", lambda r: f'{CFG["mondo"]} — {CFG["disease"]}'),
        ("Phenotypes — HPO IDs", "paste", "hpo_ids"),
        ("HPO descriptors (review)", "review", lambda r: hpo_desc(r.get("hpo_ids", ""))),
        ("Phenotypes — free text", "paste", "hpo_free"), ("NOT Phenotypes — HPO IDs", "paste", "not_hpo"),
        ("NOT Phenotypes — free text", "paste", "not_free"), ("Sex", "paste", "sex"),
        ("Country of Origin", "paste", "country"), ("Ethnicity", "paste", "ethnicity"), ("Race", "paste", "race"),
        ("Age Type", "paste", "age_type"), ("Age Value", "paste", "age_val"), ("Age Unit", "paste", "age_unit"),
        ("Previous Testing", "paste", "prev_testing"), ("Genome-wide method?", "paste", "genomewide"),
        ("Genotyping Method 1", "paste", "method1"), ("Genotyping Method 2", "paste", "method2"),
        ("Description of genotyping method", "paste", "geno_desc"), ("In trans?", "paste", "intrans"),
        ("Homozygous?", "paste", "homozygous"), ("Variant 1 (HGVS)", "paste", "var1"),
        ("Variant 1 ClinVar ID", "paste", "var1_clinvar"), ("Variant 2 (HGVS)", "paste", "var2"),
        ("Variant 2 ClinVar ID", "paste", "var2_clinvar"), ("Additional Information", "paste", "addl"),
        ("PMID(s)", "paste", "pmid"), ("Variant 1 score (review)", "review", "v1_score"),
        ("Variant 2 score (review)", "review", "v2_score"), ("Proband score cap 3 (review)", "review", "proband_score"),
        ("Scoring rationale (review)", "review", "rationale")]
    wi = sheet(wb, "GCI Individual Entry", COLS_I, INDIVIDUALS, freeze="C2")
    for r in wi.iter_rows(min_row=2, max_row=wi.max_row):
        if "no pmid" in (r[28].value or "").lower(): r[0].fill = FLAG

    # ---- Experimental ----
    COLS_E = [
        ("Paper (review)", "review", "paper"), ("PMID(s)", "paste", "pmid"), ("Experiment type", "paste", "etype"),
        ("Sub-type / option", "paste", "subtype"), ("Experiment name", "paste", "name"),
        ("Function GO ID", "paste", "go_id"), ("Function (free text)", "paste", "go_free"),
        ("Evidence for function", "paste", "func_evidence"), ("Interacting gene (HGNC)", "paste", "interact_gene"),
        ("Interaction type", "paste", "interact_type"), ("Detection method", "paste", "detect_method"),
        ("Interacting gene implicated in disease?", "paste", "implicated"), ("Organ/tissue Uberon ID", "paste", "uberon_id"),
        ("Organ/tissue (free text)", "paste", "uberon_free"), ("Normally expressed in tissue?", "paste", "normally_expressed"),
        ("Cells: patient / non-patient", "paste", "cell_type"), ("Normal function GO ID", "paste", "normal_func_go"),
        ("Description of gene alteration", "paste", "gene_alt"), ("Evidence for altered function", "paste", "altered_evidence"),
        ("Model / phenotype-to-rescue (HPO-MP IDs)", "paste", "model_hpo"),
        ("Model / phenotype-to-rescue (free text)", "paste", "model_free"),
        ("Human phenotype — Model (HPO IDs)", "paste", "human_hpo"), ("Human phenotype — Model (free text)", "paste", "human_free"),
        ("Method used to rescue", "paste", "rescue_method"), ("WT rescues?", "paste", "wt_rescues"),
        ("Patient variant rescues?", "paste", "pt_var_rescues"), ("Explanation / evidence location", "paste", "explanation"),
        ("Associated variant (ClinVar ID)", "paste", "assoc_clinvar"), ("Select status", "paste", "status"),
        ("Suggested points (review)", "review", "points"), ("Category cap (review)", "review", "category")]
    sheet(wb, "GCI Experimental Data", COLS_E, EXPERIMENTS, freeze="C2")

    # ---- Family ----
    COLS_F = [
        ("Family Label", "paste", "label"), ("Disease (MONDO — name)", "paste", "mondo"),
        ("Phenotypes — HPO IDs", "paste", "hpo_ids"), ("HPO descriptors (review)", "review", lambda r: hpo_desc(r.get("hpo_ids", ""))),
        ("Phenotypes — free text", "paste", "hpo_free"), ("NOT Phenotypes — HPO IDs", "paste", "not_hpo"),
        ("NOT Phenotypes — free text", "paste", "not_free"), ("Country of Origin", "paste", "country"),
        ("Ethnicity", "paste", "ethnicity"), ("Race", "paste", "race"), ("Previous Testing", "paste", "prev_testing"),
        ("Genotyping Method 1", "paste", "method1"), ("Genotyping Method 2", "paste", "method2"),
        ("Description of genotyping method", "paste", "geno_desc"), ("# AFFECTED with genotype", "paste", "n_affected"),
        ("# UNAFFECTED without biallelic genotype", "paste", "n_unaffected"), ("# segregations (AD/XL)", "paste", "n_seg"),
        ("Inconsistent segregations?", "paste", "inconsistent"), ("Consanguineous?", "paste", "consanguineous"),
        ("Pedigree location", "paste", "pedigree"), ("Published LOD?", "paste", "pub_lod"),
        ("Include LOD in aggregate?", "paste", "include_lod"), ("Proband label (link)", "paste", "proband"),
        ("Additional Information", "paste", "addl"), ("PMID(s)", "paste", "pmid"),
        ("Estimated LOD (review)", "review", "elod"), ("Notes (review)", "review", "notes")]
    sheet(wb, "GCI Family", COLS_F, FAMILIES, freeze="B2")

    # ---- Group ----
    COLS_G = [
        ("Group Label", "paste", "label"), ("Disease (MONDO)", "paste", "mondo"),
        ("Phenotypes — HPO IDs", "paste", "hpo_ids"), ("Phenotypes — free text", "paste", "hpo_free"),
        ("NOT Phenotypes — HPO IDs", "paste", "not_hpo"), ("NOT Phenotypes — free text", "paste", "not_free"),
        ("# males", "paste", "n_male"), ("# females", "paste", "n_female"), ("Country of Origin", "paste", "country"),
        ("Ethnicity", "paste", "ethnicity"), ("Race", "paste", "race"), ("Age Type", "paste", "age_type"),
        ("Age Value (range)", "paste", "age_range"), ("Age Unit", "paste", "age_unit"),
        ("Total individuals", "paste", "total"), ("# with family info", "paste", "n_fam"),
        ("# without family info", "paste", "n_nofam"), ("# with variant in gene", "paste", "n_var"),
        ("# without variant in gene", "paste", "n_novar"), ("# with variant in other gene", "paste", "n_othergene"),
        ("Other genes (HGNC)", "paste", "othergenes"), ("Previous Testing", "paste", "prev_testing"),
        ("Genome-wide method?", "paste", "genomewide"), ("Genotyping Method 1", "paste", "method1"),
        ("Genotyping Method 2", "paste", "method2"), ("Description of genotyping method", "paste", "geno_desc"),
        ("Additional Information", "paste", "addl"), ("PMID(s)", "paste", "pmid")]
    sheet(wb, "GCI Group", COLS_G, GROUPS, freeze="B2")

    # ---- Case-Control ----
    COLS_CC = [
        ("Case-Control Label", "paste", "label"), ("Case Cohort Label", "paste", "case_label"),
        ("Disease(s) in common", "paste", "disease"), ("Phenotypes — HPO IDs", "paste", "hpo_ids"),
        ("Phenotypes — free text", "paste", "hpo_free"), ("NOT Phenotypes — HPO IDs", "paste", "not_hpo"),
        ("NOT Phenotypes — free text", "paste", "not_free"), ("Case # males", "paste", "case_male"),
        ("Case # females", "paste", "case_female"), ("Case Country", "paste", "case_country"),
        ("Case Ethnicity", "paste", "case_ethnicity"), ("Case Race", "paste", "case_race"),
        ("Case Age range", "paste", "case_age"), ("Case Previous testing", "paste", "case_prev"),
        ("Case Genome-wide?", "paste", "case_gw"), ("Case Method 1", "paste", "case_m1"),
        ("Case Method 2", "paste", "case_m2"), ("Case genotyping description", "paste", "case_desc"),
        ("# cases WITH variant", "paste", "case_var"), ("# all cases sequenced", "paste", "case_total"),
        ("Case frequency", "paste", "case_freq"), ("Other genes (case)", "paste", "case_othergenes"),
        ("Case PMID(s)", "paste", "case_pmid"), ("Control Cohort Label", "paste", "ctrl_label"),
        ("Control # males", "paste", "ctrl_male"), ("Control # females", "paste", "ctrl_female"),
        ("Control Country", "paste", "ctrl_country"), ("Control Ethnicity", "paste", "ctrl_ethnicity"),
        ("Control Race", "paste", "ctrl_race"), ("Control Age range", "paste", "ctrl_age"),
        ("Control Method 1", "paste", "ctrl_m1"), ("Control Method 2", "paste", "ctrl_m2"),
        ("# controls WITH variant", "paste", "ctrl_var"), ("# all controls sequenced", "paste", "ctrl_total"),
        ("Control frequency", "paste", "ctrl_freq"), ("Control PMID(s)", "paste", "ctrl_pmid"),
        ("Study type", "paste", "study_type"), ("Detection method", "paste", "cc_detect"),
        ("Test statistic", "paste", "stat_test"), ("Statistic value", "paste", "stat_val"),
        ("p-value", "paste", "pval"), ("CI lower", "paste", "ci_low"), ("CI upper", "paste", "ci_high"),
        ("Bias: matched demographics?", "paste", "bias1"), ("Bias: matched genetic ancestry?", "paste", "bias2"),
        ("Bias: equivalently evaluated?", "paste", "bias3"), ("Bias: other variables differ?", "paste", "bias4"),
        ("Comments", "paste", "comments"), ("Select status", "paste", "status"),
        ("Suggested points (review)", "review", "points")]
    sheet(wb, "GCI Case-Control", COLS_CC, CASECONTROL, freeze="B2")

    # ---- Segregation (eLOD helper) ----
    COLS_S = [("Family", "review", "family"), ("Publication", "review", "pub"), ("PMID", "review", "pmid"),
              ("Affected (genotype+)", "review", "affected"), ("Segregations (affected-1)", "review", "segregations"),
              ("Unaffected at-risk", "review", "unaffected"), ("Estimated LOD (AR)", "review", "elod"),
              ("Notes", "review", "notes")]
    sheet(wb, "Segregation (review)", COLS_S, SEGREGATION, freeze="A2")

    # ---- Summary ----
    su = wb.create_sheet("Summary")
    su.append([f'{CFG["gene"]} — {CFG["disease"]} ({CFG["mondo"]})']); su["A1"].font = Font(bold=True, size=12); su.append([])
    su.append(["Bucket", "Points", "Notes"])
    for c in su[3]: c.fill = PASTE; c.font = HF; c.border = B; c.alignment = CTR
    for row in [("Genetic evidence", f'{CFG["genetic_pts"]} / 12', "case-level + segregation + case-control"),
                ("Experimental evidence", f'{CFG["experimental_pts"]} / 6', "Function 2 + Functional alteration 2 + Models & Rescue 4"),
                ("TOTAL", f'{CFG["genetic_pts"]+CFG["experimental_pts"]} / 18', ""),
                ("CLASSIFICATION", CFG["classification"], "SOP v10.1")]:
        su.append(list(row))
    for r in su.iter_rows(min_row=4, max_row=su.max_row):
        for c in r: c.border = B; c.alignment = WRAP
        if r[0].value in ("TOTAL", "CLASSIFICATION"):
            for c in r: c.fill = TOT; c.font = Font(bold=True)
    for i, w in enumerate([30, 16, 60], 1): su.column_dimensions[get_column_letter(i)].width = w

    # ---- HPO Reference + Legend ----
    hr = wb.create_sheet("HPO Reference"); hr.append(["HPO ID", "Label"])
    for c in hr[1]: c.fill = PASTE; c.font = HF; c.border = B; c.alignment = CTR
    for hid, lab in HPO_LABELS.items(): hr.append([hid, lab])
    for i, w in enumerate([16, 55], 1): hr.column_dimensions[get_column_letter(i)].width = w
    lg = wb.create_sheet("Legend")
    for i, t in enumerate([
        "GCI copy-paste workbook — build at end of curation, before the Step-9 cross-check.",
        "",
        "NAVY header = PASTE column: maps 1:1 to a GCI curation-form field — paste onto the website.",
        "GREY header = REVIEW column: human-readable helper (brief description, HPO labels, scores,",
        "  rationale, eLOD) — for manual review only; do NOT paste into the GCI.",
        "",
        "HPO IDs are comma-separated IDs only; the adjacent 'HPO descriptors' column shows labels.",
        "Disease is the same for all — use 'Copy disease term from Gene-Disease Record' or the MONDO id.",
        "Family/Group/Case-Control tabs are empty unless that evidence type is used; delete unused tabs.",
        "Scores mirror references/scoring_reference.md (SOP v10.1); classification via scripts/classify.py.",
    ], 1):
        lg.cell(i, 1, value=t)
    lg.column_dimensions["A"].width = 110

    out = sys.argv[1] if len(sys.argv) > 1 else "GCI_workbook.xlsx"
    wb.save(out); print("saved:", out, "| tabs:", wb.sheetnames)

if __name__ == "__main__":
    build()
